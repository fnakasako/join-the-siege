"""
Document processing utilities for converting various file types to images and extracting text.
These utilities support the file type processors by handling the technical details of 
document conversion and text extraction.
"""

import io
import base64
from typing import Optional, Union
from PIL import Image, ImageDraw, ImageFont
import pytesseract
from pdf2image import convert_from_bytes
import matplotlib.pyplot as plt
import matplotlib.patches as patches
from werkzeug.datastructures import FileStorage
import openpyxl
from docx import Document
import tempfile
import os

def pdf_to_image(file: FileStorage, page: int = 0, dpi: int = 200) -> bytes:
    """
    Convert PDF page to image bytes
    
    Args:
        file: FileStorage object containing PDF
        page: Page number to convert (0-indexed)
        dpi: Resolution for conversion
        
    Returns:
        Image data as bytes
    """
    try:
        file.seek(0)
        pdf_bytes = file.read()
        file.seek(0)
        
        # Convert PDF to images
        images = convert_from_bytes(pdf_bytes, dpi=dpi, first_page=page+1, last_page=page+1)
        
        if not images:
            raise ValueError("No pages found in PDF")
        
        # Convert PIL Image to bytes
        img_buffer = io.BytesIO()
        images[0].save(img_buffer, format='PNG')
        img_buffer.seek(0)
        
        return img_buffer.getvalue()
        
    except Exception as e:
        # Return a placeholder image if conversion fails
        return _create_error_image(f"PDF conversion failed: {str(e)}")

def excel_to_image(file: FileStorage, sheet: int = 0, max_rows: int = 50, max_cols: int = 20) -> bytes:
    """
    Convert Excel sheet to image bytes
    
    Args:
        file: FileStorage object containing Excel file
        sheet: Sheet index to convert (0-indexed)
        max_rows: Maximum rows to include in image
        max_cols: Maximum columns to include in image
        
    Returns:
        Image data as bytes
    """
    try:
        file.seek(0)
        workbook = openpyxl.load_workbook(file, read_only=True)
        
        if sheet >= len(workbook.worksheets):
            sheet = 0
        
        worksheet = workbook.worksheets[sheet]
        
        # Extract data from worksheet
        data = []
        for row in worksheet.iter_rows(max_row=max_rows, max_col=max_cols, values_only=True):
            data.append([str(cell) if cell is not None else '' for cell in row])
        
        # Create matplotlib figure
        fig, ax = plt.subplots(figsize=(12, 8))
        ax.axis('tight')
        ax.axis('off')
        
        # Create table
        if data:
            table = ax.table(cellText=data, loc='center', cellLoc='left')
            table.auto_set_font_size(False)
            table.set_fontsize(8)
            table.scale(1, 1.5)
        else:
            ax.text(0.5, 0.5, 'Empty Excel Sheet', ha='center', va='center', fontsize=16)
        
        # Save to bytes
        img_buffer = io.BytesIO()
        plt.savefig(img_buffer, format='PNG', bbox_inches='tight', dpi=150)
        plt.close()
        img_buffer.seek(0)
        
        return img_buffer.getvalue()
        
    except Exception as e:
        return _create_error_image(f"Excel conversion failed: {str(e)}")

def word_to_image(file: FileStorage, page: int = 0) -> bytes:
    """
    Convert Word document to image bytes
    
    Args:
        file: FileStorage object containing Word document
        page: Page number to convert (0-indexed, approximate)
        
    Returns:
        Image data as bytes
    """
    try:
        file.seek(0)
        doc = Document(file)
        
        # Extract text from paragraphs
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        
        if not paragraphs:
            return _create_error_image("Empty Word document")
        
        # Estimate which paragraphs belong to the requested "page"
        chars_per_page = 2000  # Rough estimate
        start_char = page * chars_per_page
        
        # Find paragraphs for this page
        current_char = 0
        page_paragraphs = []
        
        for para in paragraphs:
            if current_char >= start_char:
                page_paragraphs.append(para)
                current_char += len(para)
                if current_char - start_char > chars_per_page:
                    break
            else:
                current_char += len(para)
        
        if not page_paragraphs:
            page_paragraphs = paragraphs[:5]  # Fallback to first 5 paragraphs
        
        # Create image with text
        return _create_text_image(page_paragraphs, title=f"Word Document - Page {page + 1}")
        
    except Exception as e:
        return _create_error_image(f"Word conversion failed: {str(e)}")

def optimize_image_for_llm(file: FileStorage, max_size: tuple = (1024, 1024)) -> bytes:
    """
    Optimize image for LLM processing
    
    Args:
        file: FileStorage object containing image
        max_size: Maximum dimensions (width, height)
        
    Returns:
        Optimized image data as bytes
    """
    try:
        file.seek(0)
        
        with Image.open(file) as img:
            # Convert to RGB if necessary
            if img.mode != 'RGB':
                img = img.convert('RGB')
            
            # Resize if too large
            if img.size[0] > max_size[0] or img.size[1] > max_size[1]:
                img.thumbnail(max_size, Image.Resampling.LANCZOS)
            
            # Save optimized image
            img_buffer = io.BytesIO()
            img.save(img_buffer, format='PNG', optimize=True)
            img_buffer.seek(0)
            
            return img_buffer.getvalue()
            
    except Exception as e:
        return _create_error_image(f"Image optimization failed: {str(e)}")

def ocr_extract_text(file: FileStorage) -> str:
    """
    Extract text from image using OCR
    
    Args:
        file: FileStorage object containing image
        
    Returns:
        Extracted text
    """
    try:
        file.seek(0)
        
        with Image.open(file) as img:
            # Convert to RGB if necessary
            if img.mode != 'RGB':
                img = img.convert('RGB')
            
            # Extract text using pytesseract
            text = pytesseract.image_to_string(img)
            
            return text.strip()
            
    except Exception as e:
        return f"OCR extraction failed: {str(e)}"

def extract_excel_text(file: FileStorage, sheet: int = 0, max_rows: int = 100) -> str:
    """
    Extract text content from Excel sheet
    
    Args:
        file: FileStorage object containing Excel file
        sheet: Sheet index to extract from (0-indexed)
        max_rows: Maximum rows to process
        
    Returns:
        Extracted text content
    """
    try:
        file.seek(0)
        workbook = openpyxl.load_workbook(file, read_only=True)
        
        if sheet >= len(workbook.worksheets):
            sheet = 0
        
        worksheet = workbook.worksheets[sheet]
        
        # Extract text from cells
        text_content = []
        for row in worksheet.iter_rows(max_row=max_rows, values_only=True):
            row_text = []
            for cell in row:
                if cell is not None:
                    row_text.append(str(cell))
            if row_text:
                text_content.append(' | '.join(row_text))
        
        return '\n'.join(text_content)
        
    except Exception as e:
        return f"Excel text extraction failed: {str(e)}"

def _create_error_image(error_message: str) -> bytes:
    """
    Create a simple error image with the given message
    
    Args:
        error_message: Error message to display
        
    Returns:
        Image data as bytes
    """
    # Create a simple error image
    img = Image.new('RGB', (800, 600), color='white')
    draw = ImageDraw.Draw(img)
    
    try:
        # Try to use a default font
        font = ImageFont.load_default()
    except:
        font = None
    
    # Draw error message
    draw.text((50, 250), "Document Processing Error", fill='red', font=font)
    draw.text((50, 300), error_message[:100], fill='black', font=font)
    
    # Convert to bytes
    img_buffer = io.BytesIO()
    img.save(img_buffer, format='PNG')
    img_buffer.seek(0)
    
    return img_buffer.getvalue()

def _create_text_image(paragraphs: list, title: str = "Document") -> bytes:
    """
    Create an image from text paragraphs
    
    Args:
        paragraphs: List of text paragraphs
        title: Title for the document
        
    Returns:
        Image data as bytes
    """
    # Create image
    img = Image.new('RGB', (800, 1000), color='white')
    draw = ImageDraw.Draw(img)
    
    try:
        font = ImageFont.load_default()
    except:
        font = None
    
    y_position = 50
    
    # Draw title
    draw.text((50, y_position), title, fill='blue', font=font)
    y_position += 50
    
    # Draw paragraphs
    for para in paragraphs[:10]:  # Limit to 10 paragraphs
        # Wrap text to fit image width
        words = para.split()
        lines = []
        current_line = []
        
        for word in words:
            current_line.append(word)
            line_text = ' '.join(current_line)
            if len(line_text) > 80:  # Rough character limit per line
                if len(current_line) > 1:
                    current_line.pop()
                    lines.append(' '.join(current_line))
                    current_line = [word]
                else:
                    lines.append(line_text)
                    current_line = []
        
        if current_line:
            lines.append(' '.join(current_line))
        
        # Draw lines
        for line in lines[:3]:  # Limit lines per paragraph
            if y_position > 900:  # Stop if we're near the bottom
                break
            draw.text((50, y_position), line, fill='black', font=font)
            y_position += 20
        
        y_position += 10  # Space between paragraphs
        
        if y_position > 900:
            break
    
    # Convert to bytes
    img_buffer = io.BytesIO()
    img.save(img_buffer, format='PNG')
    img_buffer.seek(0)
    
    return img_buffer.getvalue()
