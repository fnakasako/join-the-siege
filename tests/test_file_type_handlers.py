"""
Unit tests for file type handlers and document utilities
"""

import pytest
from unittest.mock import Mock, patch, MagicMock
from io import BytesIO
from PIL import Image
import openpyxl
from docx import Document

from src.classifier.file_type_handling.file_type_processors import (
    classify_file,
    classify_pdf,
    classify_image,
    classify_excel,
    classify_word
)
from src.classifier.file_type_handling.document_utils import (
    pdf_to_image,
    excel_to_image,
    word_to_image,
    optimize_image_for_llm,
    ocr_extract_text,
    extract_excel_text
)


@pytest.mark.unit
class TestFileTypeProcessors:
    """Test file type processors"""
    
    def test_classify_file_no_filename(self):
        """Test handling of file with no filename"""
        mock_file = Mock()
        mock_file.filename = None
        
        result = classify_file(mock_file)
        
        assert result['classification'] == 'no_filename'
        assert result['confidence'] == 0.0
    
    def test_classify_file_unsupported_type(self):
        """Test handling of unsupported file type"""
        mock_file = Mock()
        mock_file.filename = 'test.txt'
        
        result = classify_file(mock_file)
        
        assert result['classification'] == 'unsupported_file_type'
        assert result['confidence'] == 0.0
        assert result['metadata']['file_type'] == 'txt'
    
    @pytest.mark.parametrize("filename,expected_processor", [
        ("test.pdf", "classify_pdf"),
        ("test.jpg", "classify_image"),
        ("test.jpeg", "classify_image"),
        ("test.png", "classify_image"),
        ("test.xlsx", "classify_excel"),
        ("test.xls", "classify_excel"),
        ("test.docx", "classify_word"),
        ("test.doc", "classify_word"),
    ])
    def test_classify_file_routing(self, filename, expected_processor):
        """Test that files are routed to correct processors"""
        mock_file = Mock()
        mock_file.filename = filename
        
        with patch(f'src.classifier.file_type_handling.file_type_processors.{expected_processor}') as mock_processor:
            mock_processor.return_value = {'classification': 'test', 'confidence': 0.9}
            
            result = classify_file(mock_file)
            
            mock_processor.assert_called_once_with(mock_file, 'finance')
    
    @patch('src.classifier.file_type_handling.file_type_processors.classify_with_llm')
    @patch('src.classifier.file_type_handling.file_type_processors.pdf_to_image')
    @patch('PyPDF2.PdfReader')
    def test_classify_pdf_success(self, mock_pdf_reader, mock_pdf_to_image, mock_classify_llm):
        """Test successful PDF classification"""
        # Setup mocks
        mock_file = Mock()
        mock_file.filename = 'test.pdf'
        mock_file.seek.return_value = None
        mock_file.tell.return_value = 1024
        
        mock_reader = Mock()
        mock_reader.pages = [Mock()]
        mock_reader.pages[0].extract_text.return_value = "Sample PDF text"
        mock_pdf_reader.return_value = mock_reader
        
        mock_pdf_to_image.return_value = b"fake_image_data"
        mock_classify_llm.return_value = {
            'classification': 'invoice',
            'confidence': 0.95
        }
        
        result = classify_pdf(mock_file)
        
        assert result['classification'] == 'invoice'
        assert result['confidence'] == 0.95
        assert result['metadata']['file_type'] == 'pdf'
        assert result['metadata']['page_count'] == 1
        assert result['metadata']['has_text'] is True
    
    @patch('src.classifier.file_type_handling.file_type_processors.classify_with_llm')
    @patch('src.classifier.file_type_handling.file_type_processors.optimize_image_for_llm')
    @patch('src.classifier.file_type_handling.file_type_processors.ocr_extract_text')
    @patch('PIL.Image.open')
    def test_classify_image_success(self, mock_image_open, mock_ocr, mock_optimize, mock_classify_llm):
        """Test successful image classification"""
        # Setup mocks
        mock_file = Mock()
        mock_file.filename = 'test.jpg'
        mock_file.seek.return_value = None
        mock_file.tell.return_value = 2048
        
        mock_img = Mock()
        mock_img.width = 800
        mock_img.height = 600
        mock_img.mode = 'RGB'
        mock_img.format = 'JPEG'
        mock_image_open.return_value.__enter__.return_value = mock_img
        
        mock_optimize.return_value = b"optimized_image_data"
        mock_ocr.return_value = "Extracted text from image"
        mock_classify_llm.return_value = {
            'classification': 'drivers_license',
            'confidence': 0.88
        }
        
        result = classify_image(mock_file)
        
        assert result['classification'] == 'drivers_license'
        assert result['confidence'] == 0.88
        assert result['metadata']['file_type'] == 'image'
        assert result['metadata']['width'] == 800
        assert result['metadata']['height'] == 600
    


@pytest.mark.unit
class TestDocumentUtils:
    """Test document utility functions"""
    
    def test_create_error_image(self):
        """Test error image creation"""
        from src.classifier.file_type_handling.document_utils import _create_error_image
        
        error_data = _create_error_image("Test error message")
        
        assert isinstance(error_data, bytes)
        assert len(error_data) > 0
        
        # Verify it's a valid PNG
        img = Image.open(BytesIO(error_data))
        assert img.format == 'PNG'
    
    @patch('PIL.Image.open')
    def test_optimize_image_for_llm_success(self, mock_image_open):
        """Test image optimization"""
        mock_file = Mock()
        mock_file.seek.return_value = None
        
        # Create a mock image that's too large
        mock_img = Mock()
        mock_img.mode = 'RGB'
        mock_img.size = (2048, 1536)  # Larger than max_size
        mock_img.convert.return_value = mock_img
        mock_img.thumbnail.return_value = None
        
        # Mock the save operation
        def mock_save(buffer, **kwargs):
            buffer.write(b"fake_optimized_image_data")
        mock_img.save = mock_save
        
        mock_image_open.return_value.__enter__.return_value = mock_img
        
        result = optimize_image_for_llm(mock_file, max_size=(1024, 1024))
        
        assert isinstance(result, bytes)
        mock_img.thumbnail.assert_called_once_with((1024, 1024), Image.Resampling.LANCZOS)
    
    def test_optimize_image_for_llm_error(self):
        """Test image optimization error handling"""
        mock_file = Mock()
        mock_file.seek.side_effect = Exception("File error")
        
        result = optimize_image_for_llm(mock_file)
        
        assert isinstance(result, bytes)
        # Should return error image
        img = Image.open(BytesIO(result))
        assert img.format == 'PNG'
    
    @patch('pytesseract.image_to_string')
    @patch('PIL.Image.open')
    def test_ocr_extract_text_success(self, mock_image_open, mock_ocr):
        """Test OCR text extraction"""
        mock_file = Mock()
        mock_file.seek.return_value = None
        
        mock_img = Mock()
        mock_img.mode = 'RGB'
        mock_img.convert.return_value = mock_img
        mock_image_open.return_value.__enter__.return_value = mock_img
        
        mock_ocr.return_value = "  Extracted text from image  "
        
        result = ocr_extract_text(mock_file)
        
        assert result == "Extracted text from image"
        mock_ocr.assert_called_once_with(mock_img)
    
    @patch('openpyxl.load_workbook')
    def test_extract_excel_text_success(self, mock_load_workbook):
        """Test Excel text extraction"""
        mock_file = Mock()
        mock_file.seek.return_value = None
        
        # Mock workbook and worksheet
        mock_worksheet = Mock()
        mock_worksheet.iter_rows.return_value = [
            ('Header1', 'Header2', 'Header3'),
            ('Value1', 'Value2', 'Value3'),
            ('Data1', None, 'Data3')
        ]
        
        mock_workbook = Mock()
        mock_workbook.worksheets = [mock_worksheet]
        mock_load_workbook.return_value = mock_workbook
        
        result = extract_excel_text(mock_file)
        
        expected_lines = [
            'Header1 | Header2 | Header3',
            'Value1 | Value2 | Value3',
            'Data1 | Data3'
        ]
        assert result == '\n'.join(expected_lines)
    
    def test_extract_excel_text_error(self):
        """Test Excel text extraction error handling"""
        mock_file = Mock()
        mock_file.seek.side_effect = Exception("File error")
        
        result = extract_excel_text(mock_file)
        
        assert "Excel text extraction failed" in result


@pytest.mark.integration
class TestFileTypeIntegration:
    """Integration tests for file type handling"""
    
    def create_mock_file(self, filename: str, content: bytes = b"test content") -> Mock:
        """Helper to create mock FileStorage object"""
        mock_file = Mock()
        mock_file.filename = filename
        mock_file.read.return_value = content
        mock_file.seek.return_value = None
        mock_file.tell.return_value = len(content)
        return mock_file
    
    @patch('src.classifier.file_type_handling.file_type_processors.classify_with_llm')
    def test_end_to_end_classification_flow(self, mock_classify_llm):
        """Test complete classification flow"""
        mock_classify_llm.return_value = {
            'classification': 'invoice',
            'confidence': 0.92,
            'provider_used': 'openai_gpt4o_mini'
        }
        
        mock_file = self.create_mock_file('test_invoice.pdf')
        
        with patch('PyPDF2.PdfReader') as mock_pdf_reader:
            mock_reader = Mock()
            mock_reader.pages = [Mock()]
            mock_reader.pages[0].extract_text.return_value = "Invoice #12345"
            mock_pdf_reader.return_value = mock_reader
            
            with patch('src.classifier.file_type_handling.file_type_processors.pdf_to_image') as mock_pdf_to_image:
                mock_pdf_to_image.return_value = b"fake_image_data"
                
                result = classify_file(mock_file, industry='finance')
        
        assert result['classification'] == 'invoice'
        assert result['confidence'] == 0.92
        assert result['metadata']['file_type'] == 'pdf'
        assert result['industry'] == 'finance'
        
        # Verify LLM was called with correct parameters
        mock_classify_llm.assert_called_once()
        call_args = mock_classify_llm.call_args
        assert call_args[1]['industry'] == 'finance'
        assert 'Invoice #12345' in call_args[1]['text_content']
